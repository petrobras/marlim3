import openpyxl
from typing import List, Dict, Any

class VGLObject:
    def __init__(self, iId: int = -100, fCompProducao: float = 0, fCompServico: float = 0,
                 sType: str = "Orifício", fDiameter: float = 0.007938, fCv: float = 0.9,
                 fPCalib: float = 1022.0, fTCalib: float = 25.0, fAreaRatio: float = 0.262):
        self.iId = iId
        self.fCompProducao = fCompProducao
        self.fCompServico = fCompServico
        self.sType = sType
        self.fDiameter = fDiameter
        self.fCv = fCv
        self.fPCalib = fPCalib
        self.fTCalib = fTCalib
        self.fAreaRatio = fAreaRatio

        if sType == "Orifício":
            self.iType = 0
        elif sType == "Pressão":
            self.iType = 1
        elif sType == "Venturi":
            self.iType = 2
        else:
            raise ValueError("Invalid value for sType")



    def WriteVGLObjectToJson(self, oVGLCollectionNode: Dict[str, Any]):
        
        oVGLNode = {
            "id": self.iId,
            "comprimentoMedidoProducao": self.fCompProducao,
            "comprimentoMedidoServico": self.fCompServico,
            "tipoValvula": self.iType,
            "diametroOrificio": self.fDiameter,
            "cdvgl": self.fCv,
            "pressaoCalibracao": self.fPCalib,
            "temperaturaCalibracao": self.fTCalib,
            "razaoArea": self.fAreaRatio,
            "colunaEanular": True                           # Com "True", só preciso informar o comprimentoMedidoProducao!
            }
       
        oVGLCollectionNode.append(oVGLNode)


    # -----------> NOVA IMPLEMENTAÇÃO DE "ImportVGLObjectsFromFlowedit" - JUL/2026
            # Método para pré-importar VGLs da Flowedit:
    @staticmethod
    def ImportVGLObjectsFromFlowedit(sFloweditFilePath: str,
                                        worksheet=None
                                     ) -> List['VGLObject']:
        vgl_objects = []

        # Carregar o arquivo Excel, se necessário:
        if worksheet is None:
            workbook = openpyxl.load_workbook(sFloweditFilePath)
            worksheet = workbook["Modelo Marlim 3"]

        # Localizar a célula "Acessórios"
        iRow = None
        iColumn = None
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value == "Acessórios":
                    iRow = cell.row
                    iColumn = cell.column
                    break
            if iRow is not None and iColumn is not None:
                break

        if iRow is None or iColumn is None:
            raise ValueError("Célula 'Acessórios' não encontrada!")

        # Dictionary para armazenar os comprimentos:
        oCompMedidos = {}

        # Armazenar todos os comprimentos:
        for row in worksheet.iter_rows(min_row=iRow + 4, min_col=iColumn, max_col=iColumn + 2):
            cell_value = row[0].value
            if cell_value:
                oCompMedidos[cell_value] = float(row[0].offset(row=0, column=2).value)

        # Dentre todas as "posições", localizar as que são VGL:
        for key in oCompMedidos.keys():
            if key.startswith("VGL"):
                vgl_index = int(key[3:])
                fCompProducao = oCompMedidos[key]

                # Encontrar o MGL correspondente:
                mgl_key = "MGL" + str(vgl_index)
                fCompServico = oCompMedidos.get(mgl_key, 0)  # Valor default de 0 caso MGLi não seja encontrada

                vgl_objects.append(VGLObject(iId=(vgl_index-1), fCompProducao=fCompProducao, fCompServico=fCompServico))

        return vgl_objects
    # -----------> FIM DA NOVA IMPLEMENTAÇÃO DE "ImportVGLObjectsFromFlowedit" - JUL/2026

    # -----------> IMPLEMENTAÇÃO ORIGINAL DE "ImportVGLObjectsFromFlowedit" - COMENTADA EM JUL/2026
    # (para voltar a usar, basta descomentar CADA LINHA do bloco comentado abaixo)
    #    # Método para pré-importar VGLs da Flowedit:
    #@staticmethod
    #def ImportVGLObjectsFromFlowedit(sFloweditFilePath: str) -> List['VGLObject']:
    #    vgl_objects = []
    #
    #    # Carregar o arquivo Excel:
    #    workbook = openpyxl.load_workbook(sFloweditFilePath)
    #    worksheet = workbook["Modelo Marlim 3"]
    #
    #    # Localizar a célula "Posições"
    #    iRow = None
    #    iColumn = None
    #    for row in worksheet.iter_rows():
    #        for cell in row:
    #            if cell.value == "Posições":
    #                iRow = cell.row
    #                iColumn = cell.column
    #                break
    #        if iRow is not None and iColumn is not None:
    #            break
    #
    #    if iRow is None or iColumn is None:
    #        raise ValueError("Célula 'Posições' não encontrada!")
    #
    #    # Dictionary para armazenar os comprimentos:
    #    oCompMedidos = {}
    #
    #    # Armazenar todos os comprimentos:
    #    for row in worksheet.iter_rows(min_row=iRow + 4, min_col=iColumn, max_col=iColumn + 2):
    #        cell_value = row[0].value
    #        if cell_value:
    #            oCompMedidos[cell_value] = float(row[0].offset(row=0, column=2).value)
    #
    #    # Dentre todas as "posições", localizar as que são VGL:
    #    for key in oCompMedidos.keys():
    #        if key.startswith("VGL"):
    #            vgl_index = int(key[3:])
    #            fCompProducao = oCompMedidos[key]
    #
    #            # Encontrar o MGL correspondente:
    #            mgl_key = "MGL" + str(vgl_index)
    #            fCompServico = oCompMedidos.get(mgl_key, -100)  # Valor default de -100 caso MGLi não seja encontrada
    #
    #            vgl_objects.append(VGLObject(fCompProducao=fCompProducao, fCompServico=fCompServico))
    #
    #    return vgl_objects
    # -----------> FIM DA IMPLEMENTAÇÃO ORIGINAL (COMENTADA EM JUL/2026) DE "ImportVGLObjectsFromFlowedit"