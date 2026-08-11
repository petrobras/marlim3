import os
import os.path      
import openpyxl
from typing import List, Dict
from .MaterialObject import MaterialObject
from .RockObject import RockObject
import json
from .WallObject import WallObject
from .WallLayerObject import WallLayerObject
from .PipeObject import PipeObject
from .CrossSectionObject import CrossSectionObject
from .FloweditConversionInputInfo import FloweditConversionInputInfo
from .VGLObject import VGLObject



class FloweditConversion:


    # ============> CONSTRUCTOR
    def __init__(self, oFloweditConversionInputInfo: FloweditConversionInputInfo):

        # Armazenando dados de entrada
        self.sFloweditFileToConvert = oFloweditConversionInputInfo.sFloweditFileToConvert
        self.sJsonTemplate = oFloweditConversionInputInfo.sJsonTemplate
        self.bNotWriteFluidModelToJson = oFloweditConversionInputInfo.bNotWriteFluidModelToJson
        self.bFluidModelBlackOil = oFloweditConversionInputInfo.bFluidModelBlackOil
        self.bFluidModelTab = oFloweditConversionInputInfo.bFluidModelTab
        self.bFluidModelCompositional = oFloweditConversionInputInfo.bFluidModelCompositional
        self.sFluidModelFilePath = oFloweditConversionInputInfo.sFluidModelFilePath
        self.bWriteLog = oFloweditConversionInputInfo.bWriteLog
        self.bKeepFloweditInitialXY = oFloweditConversionInputInfo.bKeepFloweditInitialXY

        # Inicializando
        self.oMaterialFloweditTableLines = None
        self.oRocksFloweditTableLines = None
        self.oCrossSectionsFloweditTableLines = None
        self.oProdPipesFloweditTableLines = None
        self.oServicePipesFloweditTableLines = None

        self.oDuctServiceFloweditTableLines = None
        self.oAnnularFloweditTableLines = None

        self.oMaterialObjectInstances = None
        self.oRockObjectInstances = None
        self.oCrossSectionObjects = None
        self.oProdPipeObjects = None
        self.oServicePipeObjects = None

        self.oInjectionDuctObjects = None
        self.oAnnularDuctObjects = None

        self.sStatusMessage = "-----"

        self.sLogFilePath = os.path.join(
                os.path.dirname(self.sFloweditFileToConvert), 
                "Log_ConversaoFLOWedit_MARLIM3.txt"
            )
        self.oLogFile = None

        # Fazendo a conversão
        self.__Convert()

    # ============> PRIVATE METHODS

        # Método centralizador da conversão:
    def __Convert(self):

        try:
            # Criar arquivo de "log":
            self.__CreateLogFile()

            # Obter a planilha onde estão as informações
            worksheet = self.__OpenFloweditFile()

            # Obter as linhas da tabela de materiais (e escrever no "log" para conferência):
            self.oMaterialFloweditTableLines = self.__GetTableLinesFromWorksheet("Materiais", 4, worksheet)
            self.__WriteTableLinesToLogFile("Materiais", self.oMaterialFloweditTableLines)

            # Obter as linhas da tabela de rochas (e escrever no "log" para conferência):
            self.oRocksFloweditTableLines = self.__GetTableLinesFromWorksheet("Rochas", 4, worksheet)
            self.__WriteTableLinesToLogFile("Rochas", self.oRocksFloweditTableLines)

            # Instanciar as rochas (e escrever no "log"):
            self.oRockObjectInstances = self.__CreateRockObjects(self.oRocksFloweditTableLines)
            self.__PrintRockObjects(self.oRockObjectInstances)

            # Instanciar os materiais:
            self.oMaterialObjectInstances = self.__CreateMaterialObjectsFromFlowedit(self.oMaterialFloweditTableLines)

            # Se for o caso, também instanciar materiais a partir das rochas:
            if self.oRockObjectInstances and self.__shouldIncludeRocksInMaterialList():
                self.__CreateMaterialObjectsFromRocks(self.oRockObjectInstances)

            # Escrever os materiais no "log" para conferência:
            self.__PrintMaterialObjects(self.oMaterialObjectInstances)

            # Escrever os materiais no json:
            self.__UpdateJsonMaterialNode(self.sJsonTemplate, self.oMaterialObjectInstances)

            # Escrever as rochas no json:
            self.__UpdateJsonFormationProperties(self.sJsonTemplate, self.oRockObjectInstances)

            # Obter as linhas da tabela de Seções Transversais (e escrever no "log" para conferência):
            self.oCrossSectionsFloweditTableLines = self.__GetTableLinesFromWorksheet("Seções transversais", 4, worksheet)
            self.__WriteTableLinesToLogFile("Seções transversais", self.oCrossSectionsFloweditTableLines)

            # Instanciar as seções transversais, e escrevê-las no json:
            self.oCrossSectionObjects = self.CreateCrossSectionInstancesFromFloweditTable(self.oCrossSectionsFloweditTableLines)
            self.WriteCrossSectionsToJson(self.oCrossSectionObjects, self.sJsonTemplate)

            # Obter as linhas da tabela de "Dutos de produção", e escrever no "log" para conferência:
            self.oProdPipesFloweditTableLines = self.__GetTableLinesFromWorksheet("Dutos de produção (início no fundo do poço)", 4, worksheet)
            self.__WriteTableLinesToLogFile("Dutos de produção (início no fundo do poço)", self.oProdPipesFloweditTableLines)

            # Instanciar os dutos de produção:
            self.oProdPipeObjects = self.CreatePipeInstancesFromFloweditTable(self.oProdPipesFloweditTableLines, "Produção", 
                                            self.oCrossSectionObjects, self.bKeepFloweditInitialXY)

            # Verificar se existe tabela de "Dutos de serviço" na FLOWedit (e caso exista, ler e instanciar os objetos):
            self.oServicePipeObjects = None
            bHasServicePipes = any(cell.value == "Dutos de serviço (início na UEP)" for row in worksheet.iter_rows() for cell in row)

            if bHasServicePipes:
                self.oServicePipesFloweditTableLines = self.__GetTableLinesFromWorksheet("Dutos de serviço (início na UEP)", 4, worksheet)
                self.oServicePipeObjects = self.CreatePipeInstancesFromFloweditTable(self.oServicePipesFloweditTableLines, "Serviço", 
                                            self.oCrossSectionObjects, self.bKeepFloweditInitialXY)

            # Escrever os dutos de produção e serviço no json:
            self.WritePipeObjectsDataToJson(self.oProdPipeObjects, self.sJsonTemplate, "dutosProducao")
            self.WritePipeObjectsDataToJson(self.oServicePipeObjects, self.sJsonTemplate, "dutosServico")

            # Obter os dados de possíveis VGLs e escrevê-los no json:
            self.ImportAndWriteVGLObjectsDataToJson(self.sJsonTemplate, worksheet)

            # Escrever as opções de modelo de fluido:
            self.WriteFluidModelDataToJson(self.sJsonTemplate)

            # Ajustar mensagem:
            self.sStatusMessage = "\nConversão realizada com sucesso!\n"
            self.__WriteToLog(self.sStatusMessage, bAlsoPrintToConsole=False)

        except Exception as e:
            self.sStatusMessage = f"Erro na conversão: {str(e)}"
            self.__WriteToLog(f"ERRO: {self.sStatusMessage}", bAlsoPrintToConsole=False)
            raise
    
        finally:

            # SEMPRE fechar o arquivo de log, mesmo em caso de erro
            self.__CloseLogFile()


        # Método para escrever as opções de modelo de fluido do usuário no json
    def WriteFluidModelDataToJson(self, sJsonFilePath: str):
        if self.bNotWriteFluidModelToJson:
            return
        
        with open(sJsonFilePath, 'r+', encoding='utf-8') as file:
            json_data = json.load(file)
            
        configuracao_inicial_node = json_data.get("configuracaoInicial", {})
        #modeloFluidoTabelaFlash_node = configuracao_inicial_node.get("modeloFluidoTabelaFlash")
        #modeloFluidoComposicional_node = configuracao_inicial_node.get("modeloFluidoComposicional")
        #pvtsimArq_node = configuracao_inicial_node.get("pvtsimArq")
        
        if self.bFluidModelBlackOil:
            #modeloFluidoTabelaFlash_node = False
            #modeloFluidoComposicional_node = False
            configuracao_inicial_node["modeloFluidoTabelaFlash"] = False
            configuracao_inicial_node["modeloFluidoComposicional"] = False
            
        if self.bFluidModelTab:
            #modeloFluidoTabelaFlash_node = True
            #modeloFluidoComposicional_node = False
            #pvtsimArq_node = os.path.basename(self.sFluidModelFilePath)
            configuracao_inicial_node["modeloFluidoTabelaFlash"] = True
            configuracao_inicial_node["modeloFluidoComposicional"] = False
            configuracao_inicial_node["pvtsimArq"] = os.path.basename(self.sFluidModelFilePath)

        if self.bFluidModelCompositional:
            #modeloFluidoTabelaFlash_node = False
            #modeloFluidoComposicional_node = True
            #pvtsimArq_node = os.path.basename(self.sFluidModelFilePath)
            configuracao_inicial_node["modeloFluidoTabelaFlash"] = False
            configuracao_inicial_node["modeloFluidoComposicional"] = True
            configuracao_inicial_node["pvtsimArq"] = os.path.basename(self.sFluidModelFilePath)
            
         # Escrever os dados modificados de volta no arquivo JSON:
        with open(sJsonFilePath, 'w', encoding='utf-8') as file:
            json.dump(json_data, file, indent=2, ensure_ascii=False)
        file.close()
            
            

        # Método para escrever as seções transversais no arquivo json:
    def WriteCrossSectionsToJson(self, oCrossSectionsToWrite: List[CrossSectionObject], oJsonFilePath: str):
        # Abrir o arquivo JSON:
        with open(oJsonFilePath, 'r+', encoding='utf-8') as file:       # acrescentei o encoding posteriormente por causa do encoding
            json_data = json.load(file)

        # Localizar o nó "secaoTransversal"
        secao_transversal_node = json_data.get('secaoTransversal', {})

        # Apagar todo o seu conteúdo:
        secao_transversal_node.clear()

        # Criar uma nova estrutura para cada instância CrossSectionObject e adicioná-la ao JSON:
        for cross_section in oCrossSectionsToWrite:
            cross_section_node = {
                "id": cross_section.iId,
                "rotulo": cross_section.sCrossSectionLabel,
                "anular": cross_section.bAnular,
                "diametroExterno": cross_section.dExternalDiameter,
                "diametroInterno": cross_section.dInternalDiameter,
                "rugosidade": cross_section.dRoughness,
                "camadas": []
            }

            # Adicionar os objetos WallLayerObjects às "camadas" em ordem ascendente das chaves "integer":
            for layer_index, wall_layer_object in sorted(cross_section.oCrossSectionLayerList.items()):
                layer_node = {
                    "tipoMedicaoCamada": wall_layer_object.sTypeLayerMeasurement,
                    "espessura": wall_layer_object.dWidthInMeters,
                    "idMaterial": wall_layer_object.idMaterial,
                    "discretizacao": wall_layer_object.iDiscretization
                }
                cross_section_node["camadas"].append(layer_node)

            # Adicionar o nó do atual CrossSectionObject ao nó "secaoTransversal" do JSON:
            #secao_transversal_node[str(cross_section.iId)] = cross_section_node
            secao_transversal_node.append(cross_section_node)

        # Escrever os dados modificados de volta no arquivo JSON:
        with open(oJsonFilePath, 'w', encoding='utf-8') as file:        # acrescentei o encoding posteriormente por causa do encoding
            json.dump(json_data, file, indent=2, ensure_ascii=False)    # acrescentei o ensure_ascii posteriormente por causa do encoding

        # Fechar o arquivo:
        file.close()




        # Método para obter da FLOWedit e escrever no json os dados de eventuais VGLs:
    def ImportAndWriteVGLObjectsDataToJson(self, sJsonFilePath: str, oFloweditWorksheet):

        # Puxar as possíveis VGLs da FLOWedit:
        oImportedVGLObjects = VGLObject.ImportVGLObjectsFromFlowedit("", oFloweditWorksheet)
        
        # Abrir o JSON:
        with open(sJsonFilePath, 'r+', encoding='utf-8') as file:
            json_data = json.load(file)

            # Encontrar (ou criar) o node sob o qual escrever as VGLs:
            oVGLCollectionNode = json_data.get("fonteGasLift")
            if oVGLCollectionNode is None:
                oVGLCollectionNode = []
                json_data["fonteGasLift"] = oVGLCollectionNode
            else:
                oVGLCollectionNode.clear()

            # Para cada VGL importada...
            if (oImportedVGLObjects is not None):
                if len(oImportedVGLObjects) > 0:
                    for oVGLObject in oImportedVGLObjects:
                        oVGLObject.WriteVGLObjectToJson(oVGLCollectionNode)

            # Salvar as modificações no JSON:
            file.seek(0)
            json.dump(json_data, file, indent=4, ensure_ascii=False)
            file.truncate()



        # Método para escrever os dados dos dutos instanciados no json do Marlim 3
    def WritePipeObjectsDataToJson(self,
                                   oPipeObjectList: List[PipeObject],
                                   sJsonFilePath: str,
                                   sJsonFileNodeToWriteUnder: str 
                                   ):

        # Indicar se existe lista de tubos instanciados:
        bPipeObjectListIsNone = (oPipeObjectList is None)
        
        # Abrir o JSON 
        with open(sJsonFilePath, 'r+', encoding='utf-8') as file:       # acrescentei o encoding posteriormente por causa do encoding
            # Carregar os dados:
            json_data = json.load(file)

            # Obter referência ao "configuracaoInicial" do json:
            initialConfigNode = json_data.get("configuracaoInicial", {})

            # Verificar se existe linha de serviço e ligar o "linhaGas", se for o caso:
            if (sJsonFileNodeToWriteUnder == "dutosServico"):
                initialConfigNode["linhaGas"] = not bPipeObjectListIsNone

            # Localizar o nó especificado:
            json_node_to_write_under = json_data.get(sJsonFileNodeToWriteUnder, {})

            # Limpar o nó especificado:
            json_node_to_write_under.clear()

            # Escrever os dados, se for o caso:
            if (not bPipeObjectListIsNone):

                # Dados de cada duto:
                for pipe_object in oPipeObjectList:
                    #if not pipe_object.bIncludeInJson:
                    #    continue
                    pipe_object.WriteDuctDataToJson(json_node_to_write_under)

                # Dados "xy" relacionados em "configuracaoInicial":
                first_pipe = oPipeObjectList[0]
                dXInicio = first_pipe.dPreviousX
                dYInicio = first_pipe.dPreviousY

                initialConfigNode["modoXY"] = True
                if (sJsonFileNodeToWriteUnder == "dutosProducao"):
                    initialConfigNode["xProdInicio"] = dXInicio
                    initialConfigNode["yProdInicio"] = dYInicio
                else:
                    if (sJsonFileNodeToWriteUnder == "dutosServico"):
                        initialConfigNode["xServInicio"] = dXInicio
                        initialConfigNode["yServInicio"] = dYInicio

            # Salvar as modificações no JSON:
            file.seek(0)
            json.dump(json_data, file, indent=4, ensure_ascii=False)    # adicionei o ensure_ascii posteriormente por causa do encoding
            file.truncate()


        # Método para instanciar objetos representando as seções transversais listadas na FLOWedit:
    def CreateCrossSectionInstancesFromFloweditTable(self, 
                                            oCrossSectionExcelTableLines: List[List[str]]) -> List[CrossSectionObject]:

        cross_section_objects = []

        oMaterialObjectList = self.oMaterialObjectInstances
        oRockObjectList = self.oRockObjectInstances

        for oExcelLine in oCrossSectionExcelTableLines[0:]:                         # Para cada seção transversal na FLOWedit...
            cross_section_instance = CrossSectionObject(oExcelLine=oExcelLine,      # ... instanciar um CrossSectionObject...
            oMaterialObjectList = oMaterialObjectList,
            oRockObjectList = oRockObjectList)

            cross_section_objects.append(cross_section_instance)                    # ... adicionar à lista de saída...

        return cross_section_objects                                                # ... e retornar a lista.





        # Método para instanciar objetos representando os tubos de dada tabela da Flowedit:
    def CreatePipeInstancesFromFloweditTable(self,
                                             oPipeExcelTableLines: List[List[str]],
                                             sPipeLabelJsonSuffix: str,
                                             oCrossSectionObjectInstances: List[CrossSectionObject],
                                             bKeepFloweditInitialXY: bool
                                             ) -> List[PipeObject]:

        # "Offset" adequado nas coordenadas "x" e "y" de acordo com opção pelo ponto inicial:
        dOffsetInX = 0.0
        dOffsetInY = 0.0
        if (not bKeepFloweditInitialXY):
            oInitialPointExcelLine = oPipeExcelTableLines[0]
            dOffsetInX = float(oInitialPointExcelLine[2])
            dOffsetInY = float(oInitialPointExcelLine[3])

        # Instanciar linha por linha:
        pipe_objects = []
        oPreviousExcelLine = oPipeExcelTableLines[0]

        for oExcelLine in oPipeExcelTableLines[1:]:  # Começando na segunda linha:
            pipe_instance = PipeObject(
                sLabelSuffix=sPipeLabelJsonSuffix,
                oExcelLine=oExcelLine,
                oPreviousExcelLine=oPreviousExcelLine,
                oCrossSectionObjectInstances=oCrossSectionObjectInstances,
                dOffsetInX=dOffsetInX,
                dOffsetInY=dOffsetInY
            )

            # Adicionar à saída:
            pipe_objects.append(pipe_instance)

            # Atualizar oPreviousExcelLine para a próxima iteração
            oPreviousExcelLine = oExcelLine

        return pipe_objects










        # Método originalmente concebido para escrever as linhas de uma tabela da Flowedit na tela acompanhadas de seus números de linha (SEM USO NO MOMENTO)
    def PrintRowIndexedTableLines(self, oExcelTableLines: Dict[int, List[str]]):
        sorted_lines = sorted(oExcelTableLines.items())

        for row_number, excel_line in sorted_lines:
            print(f"Row {row_number}: {excel_line}")


        # Método para instanciar objetos representando Paredes a partir das "Paredes" da Flowedit
    def CreateWallObjects(self, oWallsFloweditTableLines):
        wall_objects = []
        oMaterialObjectList = self.oMaterialObjectInstances
        oRockObjectList = self.oRockObjectInstances

        for index, excel_line in enumerate(oWallsFloweditTableLines):
            wall_object = WallObject(iIndex=index, oExcelLine=excel_line, oMaterialObjectList=oMaterialObjectList, oRockObjectList=oRockObjectList)
            wall_objects.append(wall_object)

        return wall_objects


        # Método para escrever as propriedades das "Rochas" da Flowedit na "Formacao" do json
    def __UpdateJsonFormationProperties(self, json_template: str, rock_objects: List[RockObject]):
        with open(json_template, 'r+', encoding='utf-8') as file:   # acrescentei o encoding depois por causa do encoding
            json_data = json.load(file)

            # Localizar "configuracaoInicial"
            configuracao_inicial_node = json_data.get("configuracaoInicial", {})

            # Verificar se "Formacao" existe:
            formacao_node = configuracao_inicial_node.get("Formacao")
            if formacao_node is None:
                # Criar "Formacao" caso não exista
                formacao_node = {
                    "TempoProducao": 365,
                    "Propriedades": []
                }
                configuracao_inicial_node["Formacao"] = formacao_node
                propriedades_node = formacao_node.get("Propriedades", []) # RodrigoGDT: adicionado posteriormente.
            else:
                # Limpar o conteúdo de "Propriedades" caso exista
                propriedades_node = formacao_node.get("Propriedades", [])
                propriedades_node.clear()

            # Criar nova estrutura para cada instância RockObject e adicionar à "Propriedades"
            for rock_object in rock_objects:
                structure = {
                    "id": rock_object.iId,
                    "rotulo": rock_object.sName,
                    "condutividade": float(rock_object.sConductivity),
                    "massaEspecifica": float(rock_object.sRho),
                    "calorEspecifico": float(rock_object.sCP)
                }
                propriedades_node.append(structure)

            # Rebobinar o arquivo e escrever os dados atualizados:
            file.seek(0)

            json.dump(json_data, file, indent=2, ensure_ascii=False)    # Linha original, acrescentei o ensure_ascii posteriormente por causa do encoding
            file.truncate()



        # Método para indicar se as rochas devem ser incluídas como materiais no json
    def __shouldIncludeRocksInMaterialList(self) -> bool:
        #return True  # Por default...
        return False  # 11-JUN-2026: Na atual versão da FLOWedit, as Rochas já vêm na tabela de Materiais

        # Método para instanciar materiais a partir de rochas quando necessário:
    def __CreateMaterialObjectsFromRocks(self, rock_objects: List[RockObject]):
        starting_index = len(self.oMaterialObjectInstances)
        for rock_object in rock_objects:
            material_object = MaterialObject(
                starting_index,
                [str(starting_index), rock_object.sName, "Sólido", rock_object.sConductivity, rock_object.sCP, rock_object.sRho, "-100", "-100"]
            )
            starting_index = starting_index + 1
            self.oMaterialObjectInstances.append(material_object)


        # Método para instanciar rochas a partir da tabela de "Rochas" da Flowedit
    def __CreateRockObjects(self, rock_lines: List[List[str]]) -> List[RockObject]:
        rock_objects = []
        for i, line in enumerate(rock_lines):
            rock_object = RockObject(i, line)
            rock_objects.append(rock_object)
        return rock_objects



        # Método para escrever os materiais no json
    def __UpdateJsonMaterialNode(self, json_template: str, material_objects: List[MaterialObject]):
        with open(json_template, 'r+', encoding='utf-8') as file:
            json_data = json.load(file)

            # Localizar o nó de materiais:
            material_node = json_data.get("material", [])

            # Apagar todo o conteúdo
            material_node.clear()

            # Criar novas estruturas para cada instãncia de MaterialObject
            for material_object in material_objects:
                structure = {
                    "id": int(material_object.iId),
                    "rotulo": material_object.sName,
                    "condutividade": float(material_object.sConductivity),
                    "calorEspecifico": float(material_object.sCP),
                    "rho": float(material_object.sRho),
                    "tipo": int(material_object.sType),
                }

                # Adicionar campo "visc" se sViscosity não for "-100"
                if material_object.sViscosity != "-100":
                    structure["visc"] = float(material_object.sViscosity)

                # Adicionar campo "coefExpTermica" se sThermalExpCoef não for "-100"
                if material_object.sThermalExpCoef != "-100":
                    structure["beta"] = float(material_object.sThermalExpCoef)

                # Acrescentar a estrutura ao nó "material"
                material_node.append(structure)

            # Rebobinar o arquivo e escrever os dados atualizados:
            file.seek(0)
            json.dump(json_data, file, indent=2, ensure_ascii=False)    # acrescentei depois o "ensureAscii" por causa do encoding
            file.truncate()


        # Método para criar o arquivo de "log":
    def __CreateLogFile(self):
        """Cria arquivo de log APENAS se bWriteLog for True.
           Se bWriteLog for False e arquivo existir, deleta o arquivo."""
    
        if self.bWriteLog:
            # Criar/abrir arquivo de log
            try:
                self.oLogFile = open(self.sLogFilePath, "w", encoding="utf-8")
                self.__WriteToLog("=== Início da Conversão FLOWedit → MARLIM 3 ===", bAlsoPrintToConsole=False)
            except IOError as e:
                print(f"Erro ao criar arquivo de log: {e}")
                self.bWriteLog = False  # Desativa log se falhar
        else:
            # Se bWriteLog é False, deletar arquivo de log anterior se existir
            if os.path.exists(self.sLogFilePath):
                try:
                    os.remove(self.sLogFilePath)
                    print(f"Arquivo de log anterior removido: {self.sLogFilePath}")
                except OSError as e:
                    print(f"Erro ao remover arquivo de log anterior: {e}")


        # Método para obter as linhas de uma tabela na Flowedit
    def __GetTableLinesFromWorksheet(self, sTableTitle: str, iRowOffset: int, worksheet) -> List[List[str]]:
        table_lines = []

        # Encontrar a célula que contém o título da tabela:
        table_start_cell = None
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value == sTableTitle:
                    table_start_cell = cell
                    break
            if table_start_cell:
                break

        if table_start_cell:
            # Determinar as fronteiras da tabela
            table_start_row = table_start_cell.row
            table_start_col = table_start_cell.column

            max_row = worksheet.max_row
            max_col = worksheet.max_column

            # Ler cabeçalhos da tabela:
            oColHeaders = []
            iColHeaderCount = 0
            for col in range(table_start_col, max_col + 1):
                header_value = worksheet.cell(row=table_start_row+iRowOffset-2, column=col).value
                if header_value is None:
                    break
                oColHeaders.append(header_value)
                iColHeaderCount = iColHeaderCount + 1

            # -----> MOSTRAR NA TELA
            #sHeadersDebug = ", ".join(str(h) for h in oColHeaders)
            #print(f"TABELA {sTableTitle}: {iColHeaderCount} colunas abaixo encontradas:")
            #print(f"  Cabeçalhos: {sHeadersDebug}")
            # -----> FIM DE EXIBIÇÃO NA TELA

            # Percorrer as linhas da tabela
            #for row in range(table_start_row, max_row + 1):
            for row in range(table_start_row+iRowOffset, max_row + 1):   # RodrigoGDT: Ler apenas o próprio conteúdo da tabela?
                line = []
                iCountColsReadForThisLine = 0
                for col in range(table_start_col, max_col + 1):
                    value = worksheet.cell(row=row, column=col).value
                    iCountColsReadForThisLine = iCountColsReadForThisLine + 1
                    #if value is None:       # --> critério original (substituído pela linha abaixo)
                    if iCountColsReadForThisLine > iColHeaderCount:
                        break  # Coluna em branco indica fim da tabela
                    line.append(value if value is not None else "")

                #if not line:       # --> critério original (substituído pela linha abaixo)
                if all(item == "" or item is None for item in line):
                    break  # linha em branco indica o fim da tabela

                table_lines.append(line)

        # =============================================
        #   MODIFICAÇÃO EM 24/7/2023 PARA APANHAR A T. AMBIENTE DE LINHAS "Início" das Geometrias
        if "Geometria" in sTableTitle:
            if len(table_lines[0]) == 3:
                if table_lines[0][0] == "Início":
                    iLastLineLength = len(table_lines[len(table_lines)-1])
                    if iLastLineLength > 3:
                        for col in range(table_start_col+3, table_start_col+iLastLineLength):
                            sCellValue = worksheet.cell(row=table_start_row+iRowOffset, column=col).value
                            if sCellValue is None:
                                table_lines[0].append("a")
                            else:
                                table_lines[0].append(sCellValue)
        #   FIM DA MODIFICAÇÃO EM 24/7/2023
        # =============================================

        return table_lines


        # Método para ler o conteúdo de determinada da tabela, indexando as linhas pelo número da linha (SEM USO POR AGORA)
    def __GetRowIndexedTableLinesFromWorksheet(self, sTableTitle: str, iRowOffset: int, worksheet) -> Dict[int, List[str]]:
        table_lines = {}

        # Encontrar a localização do título da tabela
        table_start_cell = None
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value == sTableTitle:
                    table_start_cell = cell
                    break
            if table_start_cell:
                break

        if table_start_cell:
            # Determinar os limites da tabela
            table_start_row = table_start_cell.row + 1  # Supor que a primeira linha contém cabeçalho
            table_end_row = worksheet.max_row
            table_columns = worksheet.max_column

            # Obter as linhas da tabela indexadas por número de linha:
            #for row_number in range(table_start_row, table_end_row + 1):   # RodrigoGDT: pegar apenas o conteúdo da tabela.
            for row_number in range(table_start_row+iRowOffset, table_end_row + 1):
                row_values = []
                for column in range(1, table_columns + 1):
                    cell_value = worksheet.cell(row=row_number, column=column).value
                    row_values.append(cell_value)
                table_lines[row_number] = row_values

        return table_lines










        # Método para instanciar os materiais da Flowedit
    def __CreateMaterialObjectsFromFlowedit(self, table_lines: List[List[str]]) -> List[MaterialObject]:
        material_objects = []
        for i, line in enumerate(table_lines):
            material_object = MaterialObject(i, line)
            material_objects.append(material_object)
        return material_objects



        # Método para abrir o arquivo da Flowedit e localizar a planilha com os dados:
        # IMPLEMENTAÇÃO ORIGINAL (COMENTADA):    
    #def __OpenFloweditFile(self):
    #    if os.path.exists(self.sFloweditFileToConvert):
    #        workbook = openpyxl.load_workbook(self.sFloweditFileToConvert)
    #        worksheet_name = "Modelo Marlim 3"
    #        if worksheet_name in workbook.sheetnames:
    #            return workbook[worksheet_name]
    #        else:
    #            self.__WriteToLog("Planilha Modelo Marlim 3 não encontrada no arquivo da Flowedit!")
    #            # Parar o programa 
    #            exit()
    #    else:
    #        self.__WriteToLog("Arquivo da Flowedit inexistente!")
    #        # Parar o programa
    #        exit()

        # IMPLEMENTAÇÃO ATUALIZADA:
    def __OpenFloweditFile(self):
        """Abre arquivo FLOWedit e retorna worksheet 'Modelo Marlim 3'
           Lança exceções em caso de erro."""
    
        if not os.path.exists(self.sFloweditFileToConvert):
            sErrorMsg = f"Arquivo da Flowedit inexistente: {self.sFloweditFileToConvert}"
            self.__WriteToLog(f"ERRO: {sErrorMsg}")
            raise FileNotFoundError(sErrorMsg)
    
        try:
            workbook = openpyxl.load_workbook(self.sFloweditFileToConvert)
        except Exception as e:
            sErrorMsg = f"Erro ao abrir arquivo Excel: {e}"
            self.__WriteToLog(f"ERRO: {sErrorMsg}")
            raise ValueError(sErrorMsg)
    
        worksheet_name = "Modelo Marlim 3"
        if worksheet_name not in workbook.sheetnames:
            sErrorMsg = f"Planilha '{worksheet_name}' não encontrada. Planilhas disponíveis: {workbook.sheetnames}"
            self.__WriteToLog(f"ERRO: {sErrorMsg}")
            raise ValueError(sErrorMsg)
    
        self.__WriteToLog(f"\nArquivo FLOWedit aberto com sucesso: {self.sFloweditFileToConvert}")
        return workbook[worksheet_name]



    # ============> LOG FILE METHODS

    def __WriteToLog(self, sMessage: str, bAlsoPrintToConsole: bool = False):
        """Escreve mensagem no arquivo de log e opcionalmente no console"""
        if self.bWriteLog and self.oLogFile:
            self.oLogFile.write(sMessage + "\n")
            self.oLogFile.flush()  # Força a escrita imediata
    
        if bAlsoPrintToConsole:
            print(sMessage)



    def __CloseLogFile(self):
        """Fecha o arquivo de log se estiver aberto"""
        if self.oLogFile is not None:
            try:
                self.__WriteToLog("\n=== Fim da Conversão ===", bAlsoPrintToConsole=False)
                self.oLogFile.close()
                self.oLogFile = None
            except Exception as e:
                print(f"Erro ao fechar arquivo de log: {e}")





        # Método para escrever na tela as rochas instanciadas:
    def __PrintRockObjects(self, rock_objects: List[RockObject]):
        self.__WriteToLog("\n\nLISTA DE OBJETOS-ROCHA CRIADOS:\n", bAlsoPrintToConsole=False)
        for rock_object in rock_objects:
            self.__WriteToLog("Rocha Instanciada:", bAlsoPrintToConsole=False)
            self.__WriteToLog("---Id = " + str(rock_object.iId), bAlsoPrintToConsole=False)
            self.__WriteToLog("---Name =" + str(rock_object.sName), bAlsoPrintToConsole=False)
            self.__WriteToLog("---Conductivity =" + str(rock_object.sConductivity), bAlsoPrintToConsole=False)
            self.__WriteToLog("---CP =" + str(rock_object.sCP), bAlsoPrintToConsole=False)
            self.__WriteToLog("---Rho =" + str(rock_object.sRho), bAlsoPrintToConsole=False)
            self.__WriteToLog("\n" * 1, bAlsoPrintToConsole=False)

        # Método para escrever na tela os materiais instanciados:
    def __PrintMaterialObjects(self, material_objects: List[MaterialObject]):
        self.__WriteToLog("\n\nLISTA DE OBJETOS-MATERIAL CRIADOS:\n", bAlsoPrintToConsole=False)
        for material_object in material_objects:
            self.__WriteToLog("Material Instanciado:", bAlsoPrintToConsole=False)
            self.__WriteToLog("---Id = " + str(material_object.iId), bAlsoPrintToConsole=False)
            self.__WriteToLog("---Name =" + str(material_object.sName), bAlsoPrintToConsole=False)
            self.__WriteToLog("---Type =" + str(material_object.sType), bAlsoPrintToConsole=False)
            self.__WriteToLog("---Conductivity =" + str(material_object.sConductivity), bAlsoPrintToConsole=False)
            self.__WriteToLog("---CP =" + str(material_object.sCP), bAlsoPrintToConsole=False)
            self.__WriteToLog("---Rho =" + str(material_object.sRho), bAlsoPrintToConsole=False)
            self.__WriteToLog("---Viscosity =" + str(material_object.sViscosity), bAlsoPrintToConsole=False)
            self.__WriteToLog("---ThermalExpCoef =" + str(material_object.sThermalExpCoef), bAlsoPrintToConsole=False)
            self.__WriteToLog("\n" * 1, bAlsoPrintToConsole=False)




        # Método para escrever na tela os detalhes das paredes instanciadas:
    def __PrintWallObjects(self, wall_objects: List[WallObject]):
        for wall_object in wall_objects:
            print("PAREDE INSTANCIADA:")
            print("---Id =", wall_object.iId)
            print("---Wall Name =", wall_object.sWallName)
            #print("---Wall Layer Width List =", wall_object.sWallLayerWidthList)
            #print("---Wall Layer Material List =", wall_object.sWallLayerMaterialList)
            print("---Total Wall Thickness (m) =", wall_object.dTotalWallThicknessInMeters)
            print("---Has Completion Fluid =", wall_object.bHasCompletionFluid)
            print("---Wall Ends in Rock =", wall_object.bWallEndsInRock)
            print("\n")

            for index, wall_layer in wall_object.oWallLayerList.items():
                print("------Camada", index)
                print("------Sequential Order =", wall_layer.iSequentialOrder)
                print("------Width (m) =", wall_layer.dWidthInMeters)
                print("------Id Material =", wall_layer.idMaterial)
                print("------Is Completion Fluid =", wall_layer.bIsCompletionFluid)
                print("------Is Rock Layer =", wall_layer.bIsRockLayer)
                print("------Type Layer Measurement =", wall_layer.sTypeLayerMeasurement)
                print("------Discretization =", wall_layer.iDiscretization)
                print("------Given Material Name =", wall_layer.sGivenMaterialName)
                print("\n")
        
            if wall_object.oExternalWallLayer is not None:
                print("---Camada externa da parede:")
                print("------Sequential Order =", wall_object.oExternalWallLayer.iSequentialOrder)
                print("------Width (m) =", wall_object.oExternalWallLayer.dWidthInMeters)
                print("------Id Material =", wall_object.oExternalWallLayer.idMaterial)
                print("------Is Completion Fluid =", wall_object.oExternalWallLayer.bIsCompletionFluid)
                print("------Is Rock Layer =", wall_object.oExternalWallLayer.bIsRockLayer)
                print("------Type Layer Measurement =", wall_object.oExternalWallLayer.sTypeLayerMeasurement)
                print("------Discretization =", wall_object.oExternalWallLayer.iDiscretization)
                print("------Given Material Name =", wall_object.oExternalWallLayer.sGivenMaterialName)
                print("\n")
        
            print("\n" * 2)


        # Método para escrever no arquivo de "log" as linhas de uma tabela importadas da Flowedit
    def __WriteTableLinesToLogFile(self, sTableName: str, table_lines: List[List[str]]):
        self.__WriteToLog(f"\nLINHAS DA TABELA {sTableName} OBTIDAS DA FLOWEDIT:\n")
        for line in table_lines:
            formatted_line = [str(item) for item in line]
            self.__WriteToLog(" ".join(formatted_line) + "\n")